import re
from collections import Counter
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from operator import itemgetter
from openpyxl import Workbook
import pandas as pd
#nltk.download('stopwords')
#nltk.download('punkt')

wb = Workbook()
ws = wb.active

sentences_sizes = []
nb_words = []
mentions = []
hashtags = []
emotes = []
sentiments = []
f = open("dev.txt", "r")
for line in f:
    # Recupere le tweet
    sentence = re.search(r"\t.*\t", line, flags=re.IGNORECASE).group()
    # Recupere le sentiment du tweet
    sentiment = re.sub(r"^.+\t", '', line, flags=re.IGNORECASE)
    sentiment = sentiment.rstrip()
    #retire les urls
    sentence = re.sub(r"http\S+", '', sentence, flags=re.IGNORECASE)
    # Recupere les emoticones
    emote = re.findall(r"([&][#][\w]*[;])", sentence, flags=re.IGNORECASE)
    sentence = re.sub(r"([&][#][\w]*[;])", '', sentence, flags=re.IGNORECASE)
    #retire les mentions @.....
    mention = re.findall(r"@\S+", sentence, flags=re.IGNORECASE)
    sentence = re.sub(r"@\S+", '', sentence, flags=re.IGNORECASE)
    #retire les hashtags
    hashtag = re.findall(r"#\S+", sentence, flags=re.IGNORECASE)
    sentence = re.sub(r"#\S+", '', sentence, flags=re.IGNORECASE)
    #tokenise la phrase
    sentence_tokens = word_tokenize(sentence)

    # classe les tweets par leur sentiments
    if sentiment == "negative":
        sentiment = 0
    elif sentiment == "positive":
        sentiment = 1
    elif sentiment == "mixed":
        sentiment = 2
    else:
        sentiment = 3

    sentences_sizes.append(len(sentence))
    nb_words.append(len(sentence_tokens))
    mentions.append(len(mention))
    hashtags.append(len(hashtag))
    emotes.append(len(emote))
    sentiments.append(sentiment)

def writecsv(table, colid, colname):
    ws.cell(row=1, column=colid).value = colname
    i = 2
    for elt in table:
        try:
            iterator = iter(elt)
        except TypeError:
            ws.cell(row=i, column=colid).value = elt
        else:
            ws.cell(row=i, column=colid).value = ' '.join(elt)
        i = i + 1

writecsv(sentences_sizes, 1, 'sentsize')
writecsv(nb_words, 2, 'nbwords')
writecsv(mentions, 3, 'mentions')
writecsv(hashtags, 4, 'hashtags')
writecsv(emotes, 5, 'emojis')
writecsv(sentiments, 6, 'sentiment')
wb.save("test.xlsx")
